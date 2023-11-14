import sys
from PyQt6.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout, QGridLayout, QWidget
from PyQt6.QtGui import QPixmap
from PyQt6.QtCore import Qt
from openpyxl import load_workbook
import random


class PokemonGame(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Guess the Pokémon")
        self.setGeometry(100, 100, 400, 300)
        # Load Pokémon data from Excel file
        self.pokemon_data = self.load_pokemon_data()

        # Initialize game variables
        self.current_pokemon = None
        self.score = 0
        self.initialize_game()

        # Create GUI components
        self.label_image = QLabel(self)
        self.label_name = QLabel(self)
        self.label_score = QLabel(self)
        self.buttons = [QPushButton(self) for _ in range(4)]
        self.button_next = QPushButton("Next Pokémon", self)

        # Setup layout
        layout = QVBoxLayout()

        # Add image, name, and score labels to layout
        layout.addWidget(self.label_image)
        layout.addWidget(self.label_name)
        layout.addWidget(self.label_score)

        # Create a grid layout for answer choices
        grid_layout = QGridLayout()
        for i, button in enumerate(self.buttons):
            row, col = divmod(i, 2)
            grid_layout.addWidget(button, row, col, alignment=Qt.AlignmentFlag.AlignCenter)
            grid_layout.setColumnStretch(col, 1) 
        # Add grid layout to the main layout
        layout.addLayout(grid_layout)

        # Add Next Pokémon button to layout
        layout.addWidget(self.button_next)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        # Connect button click signal to the check_answer method
        for button in self.buttons:
            button.clicked.connect(self.check_answer)

        self.button_next.clicked.connect(self.next_pokemon)

        # Show the first Pokémon
        self.next_pokemon()

    def initialize_game(self):
        # Shuffle the Pokémon data
        random.shuffle(self.pokemon_data)
        # Set the index to the first Pokémon
        self.current_pokemon_index = 0
        # Reset the score
        self.score = 0

    def load_pokemon_data(self):
        # Load Pokémon data from the Excel file
        workbook = load_workbook("pokemon_data.xlsx")
        sheet = workbook.active

        # Assuming the Excel file has two columns: Pokémon names and image file paths
        pokemon_data = [{"name": cell[0].value, "image_path": cell[1].value} for cell in sheet.iter_rows(min_row=2)]

        return pokemon_data

    def next_pokemon(self):
        # Check if the player won the game
        if self.score >= 10:
            self.label_score.setText(f"Congratulations! You've won the game with a score of {self.score}!")
        else:
            self.label_score.setText(f"Final Score: {self.score}")
        # Check if there are more Pokémon to guess
        if self.current_pokemon_index < len(self.pokemon_data):
            # Get the next Pokémon data
            self.current_pokemon = self.pokemon_data[self.current_pokemon_index]

            # Display Pokémon image
            #pixmap = QPixmap(self.current_pokemon["image_path"])
            pixmap = QPixmap(self.current_pokemon["image_path"]).scaled(120, 120, aspectRatioMode=Qt.AspectRatioMode.IgnoreAspectRatio, transformMode=Qt.TransformationMode.SmoothTransformation)

            self.label_image.setPixmap(pixmap)
            self.label_image.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.label_image.setScaledContents(False)

            # Display Pokémon name
            self.label_name.setText("Who's that Pokémon?")

            # Display score
            self.label_score.setText(f"Score: {self.score}")

            # Set up answer choices
            choices_text = random.sample([p["name"] for p in self.pokemon_data if p["name"] != self.current_pokemon["name"]], 3)
            choices_text.append(self.current_pokemon["name"])
            random.shuffle(choices_text)

            for i, button in enumerate(self.buttons):
                button.setText(choices_text[i])

            # Enable buttons
            for button in self.buttons:
                button.setEnabled(True)

            # Move to the next Pokémon
            self.current_pokemon_index += 1
        else:
            # If all Pokémon have been guessed, show a message
            self.label_image.clear()
            self.label_name.setText("You've guessed all the Pokémon!")


    def check_answer(self):
        sender = self.sender()

        # Disable buttons to prevent multiple clicks
        for button in self.buttons:
            button.setEnabled(False)

        # Check if the selected answer is correct
        if sender.text() == self.current_pokemon["name"]:
            self.score += 1
            self.label_name.setText("Correct! Well done!")
        else:
            self.label_name.setText(f"Wrong! The correct answer is {self.current_pokemon['name']}.")

        # Display updated score
        self.label_score.setText(f"Score: {self.score}")

    def closeEvent(self, event):
        # Cleanup and close the application
        QApplication.quit()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    game = PokemonGame()
    game.show()
    sys.exit(app.exec())
